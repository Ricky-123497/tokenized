from pathlib import Path
import re
import unicodedata
import os

import pandas as pd
import spacy

from indicnlp.tokenize import sentence_tokenize
from indicnlp.tokenize import indic_tokenize

from openpyxl import load_workbook
from openpyxl.styles import Alignment


# =====================================================
# CONFIG
# =====================================================

DATA_DIR = Path("data")
OUTPUT_DIR = Path("output")

MEITEI_SENTENCE_DELIM = r"[꯫]"


# =====================================================
# LOAD SPACY
# =====================================================

nlp = spacy.load("en_core_web_sm")


# =====================================================
# TEXT CLEANING
# =====================================================

def clean_json_text(text):
    text = text.strip()

    text = re.sub(r'^\{"format":"html","output":"', '', text)
    text = re.sub(r'"\}$', '', text)

    text = text.replace("\\n", " ")
    text = text.replace("\\r", " ")
    text = text.replace("\\t", " ")

    return text


def normalize_text(text):
    text = unicodedata.normalize("NFC", text)

    text = clean_json_text(text)

    text = text.replace("\ufeff", "")
    text = text.replace("\r\n", "\n")
    text = text.replace("\r", "\n")

    text = re.sub(r"\s+", " ", text)

    return text.strip()


# =====================================================
# ENGLISH SENTENCE SPLITTING (spaCy)
# =====================================================

def split_english_sentences(text):
    text = normalize_text(text)

    if not text:
        return []

    doc = nlp(text)

    sentences = []

    for sent in doc.sents:
        s = sent.text.strip()

        if len(s) > 3:
            sentences.append(s)

    return sentences


# =====================================================
# MEITEI SENTENCE SPLITTING (IndicNLP)
# =====================================================

def split_meitei_sentences(text):
    text = normalize_text(text)

    if not text:
        return []

    try:
        sentences = sentence_tokenize.sentence_split(
            text,
            lang="hi",
            delim_pat=MEITEI_SENTENCE_DELIM
        )

    except Exception:
        sentences = re.findall(r"[^꯫]+꯫", text)

    cleaned = []

    for sentence in sentences:
        sentence = sentence.strip()

        if sentence:
            cleaned.append(sentence)

    return cleaned


# =====================================================
# TOKENIZATION
# =====================================================

def tokenize_meitei_sentence(sentence):
    tokens = indic_tokenize.trivial_tokenize_indic(sentence)
    return " ".join(tokens)


def tokenize_english_sentence(sentence):
    return sentence


# =====================================================
# FIND FILES
# =====================================================

def find_language_file(author_folder, language):

    txt_files = list(author_folder.glob("*.txt"))

    if language == "english":
        keywords = ["english", "eng", "en"]

    else:
        keywords = ["manipuri", "meitei", "mni", "mm", "mtei"]

    for txt_file in txt_files:
        name = txt_file.stem.lower()

        for keyword in keywords:
            if keyword in name:
                return txt_file

    return None


# =====================================================
# READ FILE
# =====================================================

def read_text_file(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


# =====================================================
# PROCESS AUTHOR
# =====================================================

def process_author_folder(author_folder):

    print(f"\nProcessing: {author_folder.name}")

    english_file = find_language_file(author_folder, "english")
    meitei_file = find_language_file(author_folder, "meitei")

    if english_file is None:
        print("English file not found.")
        return []

    if meitei_file is None:
        print("Manipuri file not found.")
        return []

    print("English:", english_file.name)
    print("Manipuri:", meitei_file.name)

    english_text = read_text_file(english_file)
    meitei_text = read_text_file(meitei_file)

    english_sentences = split_english_sentences(english_text)
    meitei_sentences = split_meitei_sentences(meitei_text)

    english_sentences = [
        tokenize_english_sentence(s)
        for s in english_sentences
    ]

    meitei_sentences = [
        tokenize_meitei_sentence(s)
        for s in meitei_sentences
    ]

    max_len = max(
        len(english_sentences),
        len(meitei_sentences)
    )

    rows = []

    for i in range(max_len):

        eng = english_sentences[i] if i < len(english_sentences) else ""
        mni = meitei_sentences[i] if i < len(meitei_sentences) else ""

        rows.append({
            "English Sentence": eng,
            "Meitei Sentence": mni
        })

    print("English sentences:", len(english_sentences))
    print("Meitei sentences:", len(meitei_sentences))

    return rows


# =====================================================
# FORMAT EXCEL
# =====================================================

def format_excel(excel_file):

    wb = load_workbook(excel_file)
    ws = wb.active

    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 110

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 60

    wb.save(excel_file)


# =====================================================
# MAIN
# =====================================================

def main():

    print("Current folder:", os.getcwd())

    if not DATA_DIR.exists():
        raise FileNotFoundError(
            "Create:\n"
            "data/Author/English.txt\n"
            "data/Author/Manipuri.txt"
        )

    OUTPUT_DIR.mkdir(exist_ok=True)

    author_folders = [
        f for f in DATA_DIR.iterdir()
        if f.is_dir()
    ]

    all_rows = []

    for author_folder in sorted(author_folders):
        rows = process_author_folder(author_folder)
        all_rows.extend(rows)

    if not all_rows:
        print("No output generated.")
        return

    df = pd.DataFrame(all_rows)

    output_file = OUTPUT_DIR / "combined_parallel_sentences.xlsx"

    df.to_excel(
        output_file,
        index=False,
        header=False
    )

    format_excel(output_file)

    print("\nDone.")
    print("Saved:", output_file)
    print("Total rows:", len(all_rows))


if __name__ == "__main__":
    main()
